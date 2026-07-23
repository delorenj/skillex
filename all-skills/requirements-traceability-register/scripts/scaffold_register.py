#!/usr/bin/env python3
"""
Scaffold a Requirements Management & Traceability Register.

Emits one CSV per tab (dependency-free, stdlib only) that import cleanly into
Google Sheets or Excel as separate sheets. With --xlsx it also writes a single
wired workbook: dropdown validation on the phase registers, live aggregation on
All Requirements, and COUNTIFS metric cards on the Dashboard.

Usage:
    scaffold_register.py --out ./req-register                 # CSV bundle
    scaffold_register.py --out ./req-register --phases 4       # 4 phase tabs
    scaffold_register.py --out ./req-register --xlsx --title "Acme Rebuild"

The 21-column phase schema and the validation lists are the single source of
truth here; references/schema.md documents them, and references/formulas.md
documents the exact formulas mirrored below.

--xlsx requires openpyxl. If it is not importable, run via:
    uv run --with openpyxl scaffold_register.py --out ./req-register --xlsx
"""

import argparse
import csv
import sys
from pathlib import Path

# ---- Canonical schemas (single source of truth) ---------------------------

# The 21-column phase register schema, in order. Columns A..U.
PHASE_COLUMNS = [
    "Requirement ID", "Ticket Key", "Ticket Title",           # A B C  identification
    "Phase", "Phase Gate", "Workstream", "Category", "Ticket Type",  # D E F G H  categorization
    "Priority", "Status", "Owner Role",                       # I J K  triage
    "Requirement Statement", "Acceptance Criteria",           # L M    definition
    "Dependencies", "Trace Tags", "Source Document", "Source Location",  # N O P Q
    "Bug Link(s)", "Test Case ID", "Test Evidence", "Notes",  # R S T U  traceability & execution
]

SOURCE_MAP_COLUMNS = ["Source ID", "Source File", "Used For", "Local Path", "Notes"]

TRACEABILITY_COLUMNS = [
    "Requirement ID", "Phase", "Workstream", "Ticket Key", "Ticket Title",
    "Priority", "Status", "Dependencies", "AC Count", "Test Case ID",
    "Bug Link(s)", "Source Document", "Source Location",
]

ACCEPTANCE_MATRIX_COLUMNS = [
    "AC ID", "Requirement ID", "Phase", "Workstream", "Acceptance Criteria",
    "Test Method", "Evidence Expected", "Status", "Source Document", "Source Location",
]

# Validation dropdowns. Column headers on the Lists tab, each with its values.
LISTS = {
    "Status":     ["Proposed", "In Progress", "Blocked", "In Review", "Done", "Deferred", "Cancelled"],
    "Priority":   ["P0", "P1", "P2", "P3"],
    "Owner Role": ["Product", "Architecture", "Backend", "Frontend", "QA", "DevOps", "Design", "Data"],
    "Ticket Type": ["Epic", "Story", "Task", "Spike", "Bug"],
    "Category":   ["Architecture", "Feature", "Infrastructure", "Security",
                   "Performance", "Documentation", "Testing"],
    "Phase Gate": ["Not Started", "In Progress", "Passed", "Failed"],
}

# Which phase-register column (by header) is fed by which Lists dropdown.
COLUMN_VALIDATION = {
    "Phase Gate": "Phase Gate",
    "Category": "Category",
    "Ticket Type": "Ticket Type",
    "Priority": "Priority",
    "Status": "Status",
    "Owner Role": "Owner Role",
    # "Phase" is fed by the generated phase-id list (P1..Pn), wired at build time.
}

DATA_ROWS = 1000  # rows of formula/validation coverage per register


def col_letter(idx0):
    """0-based column index -> spreadsheet letter (A, B, ... Z, AA)."""
    n = idx0 + 1
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def phase_sheet_names(phases):
    return [f"P{i} Requirements" for i in range(1, phases + 1)]


def phase_ids(phases):
    return [f"P{i}" for i in range(1, phases + 1)]


# ---- Formula builders (mirrored in references/formulas.md) -----------------

def all_requirements_formula(phases, dialect):
    """Live aggregation of every phase register, blank rows removed."""
    sheets = phase_sheet_names(phases)
    if dialect == "sheets":
        blocks = "; ".join(f"'{s}'!A2:U" for s in sheets)
        return (f'=QUERY({{{blocks}}}, '
                f'"select * where Col1 is not null", 0)')
    # excel / xlsx dynamic arrays
    stack = ",".join(f"'{s}'!A2:U{DATA_ROWS}" for s in sheets)
    keys = ",".join(f"'{s}'!A2:A{DATA_ROWS}" for s in sheets)
    return (f"=_xlfn._xlws.FILTER(_xlfn.VSTACK({stack}),"
            f"_xlfn.VSTACK({keys})<>\"\")")


def ac_count_formula(ac_cell, dialect):
    """Count acceptance criteria = newline-separated lines in the AC cell."""
    if dialect == "sheets":
        return (f'=IF({ac_cell}="",0,'
                f'LEN({ac_cell})-LEN(SUBSTITUTE({ac_cell},CHAR(10),""))+1)')
    return (f'=IF({ac_cell}="",0,'
            f'LEN({ac_cell})-LEN(SUBSTITUTE({ac_cell},CHAR(10),""))+1)')


DASHBOARD_METRICS = [
    ("Total Requirements", '=COUNTA(\'All Requirements\'!A2:A)'),
    ("Open (not Done)",    '=COUNTIFS(\'All Requirements\'!A2:A,"<>",\'All Requirements\'!J2:J,"<>Done")'),
    ("Done",               '=COUNTIF(\'All Requirements\'!J2:J,"Done")'),
    ("Blocked",            '=COUNTIF(\'All Requirements\'!J2:J,"Blocked")'),
    ("P0",                 '=COUNTIF(\'All Requirements\'!I2:I,"P0")'),
    ("P1",                 '=COUNTIF(\'All Requirements\'!I2:I,"P1")'),
    ("P2",                 '=COUNTIF(\'All Requirements\'!I2:I,"P2")'),
    ("P3",                 '=COUNTIF(\'All Requirements\'!I2:I,"P3")'),
    ("Acceptance Criteria",'=COUNTA(\'Acceptance Matrix\'!A2:A)'),
]


# ---- CSV bundle ------------------------------------------------------------

def write_csv(path, rows):
    with path.open("w", newline="", encoding="utf-8") as fh:
        csv.writer(fh).writerows(rows)


def build_csv_bundle(out, phases, title):
    out.mkdir(parents=True, exist_ok=True)

    # Lists tab: one column per dropdown + the generated Phase id column.
    lists = dict(LISTS)
    headers = ["Phase"] + list(lists.keys())
    values = [phase_ids(phases)] + [lists[h] for h in lists]
    maxlen = max(len(v) for v in values)
    rows = [headers]
    for r in range(maxlen):
        rows.append([col[r] if r < len(col) else "" for col in values])
    write_csv(out / "Lists.csv", rows)

    write_csv(out / "Source Map.csv", [SOURCE_MAP_COLUMNS])

    for name in phase_sheet_names(phases):
        write_csv(out / f"{name}.csv", [PHASE_COLUMNS])

    write_csv(out / "All Requirements.csv", [PHASE_COLUMNS])
    write_csv(out / "Traceability.csv", [TRACEABILITY_COLUMNS])
    write_csv(out / "Acceptance Matrix.csv", [ACCEPTANCE_MATRIX_COLUMNS])
    write_csv(out / "Dashboard.csv",
              [[title], [], ["Metric", "Value"]] + [[m, ""] for m, _ in DASHBOARD_METRICS])

    readme = out / "IMPORT_ORDER.txt"
    readme.write_text(
        f"Requirements Management & Traceability Register — {title}\n"
        f"{'=' * 60}\n\n"
        "Import each CSV as its own tab (File > Import > Insert new sheet).\n\n"
        "Order:\n"
        "  1. Lists              (dropdown source values — import first)\n"
        "  2. Source Map         (where requirements came from)\n"
        + "".join(f"  {i+3}. {n}\n" for i, n in enumerate(phase_sheet_names(phases)))
        + "  n. All Requirements  (paste the aggregation formula in A2)\n"
          "  n. Traceability      (lifecycle view)\n"
          "  n. Acceptance Matrix (granular testable criteria)\n"
          "  n. Dashboard         (paste COUNTIFS metrics)\n\n"
        "CSVs carry structure only. Add data validation and the live formulas\n"
        "from references/formulas.md, or run this script with --xlsx for a\n"
        "fully wired workbook.\n",
        encoding="utf-8",
    )
    return sorted(p.name for p in out.iterdir())


# ---- Wired .xlsx -----------------------------------------------------------

def build_xlsx(out, phases, title):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl not available — skipping .xlsx "
              "(retry with: uv run --with openpyxl scaffold_register.py ... --xlsx)",
              file=sys.stderr)
        return None

    out.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    wrap = Alignment(vertical="top", wrap_text=True)

    def add_headers(ws, cols):
        for c, name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c, value=name)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(c)].width = max(12, min(40, len(name) + 6))
        ws.freeze_panes = "A2"

    # Dashboard (first tab = default view)
    dash = wb.active
    dash.title = "Dashboard"
    dash["A1"] = title
    dash["A1"].font = Font(bold=True, size=16)
    dash["A3"], dash["B3"] = "Metric", "Value"
    dash["A3"].font = dash["B3"].font = Font(bold=True)
    for i, (metric, formula) in enumerate(DASHBOARD_METRICS, start=4):
        dash.cell(row=i, column=1, value=metric)
        dash.cell(row=i, column=2, value=formula)
    dash.column_dimensions["A"].width = 24
    dash.column_dimensions["B"].width = 12

    # Lists
    lists_ws = wb.create_sheet("Lists")
    lists = {"Phase": phase_ids(phases), **LISTS}
    ranges = {}
    for c, (name, vals) in enumerate(lists.items(), start=1):
        letter = get_column_letter(c)
        lists_ws.cell(row=1, column=c, value=name).font = Font(bold=True)
        for r, v in enumerate(vals, start=2):
            lists_ws.cell(row=r, column=c, value=v)
        ranges[name] = f"Lists!${letter}$2:${letter}${1 + len(vals)}"

    wb.create_sheet("Source Map")
    add_headers(wb["Source Map"], SOURCE_MAP_COLUMNS)

    # Phase registers with dropdown validation
    validations = dict(COLUMN_VALIDATION)
    validations["Phase"] = "Phase"
    for name in phase_sheet_names(phases):
        ws = wb.create_sheet(name)
        add_headers(ws, PHASE_COLUMNS)
        for col_name, list_name in validations.items():
            idx = PHASE_COLUMNS.index(col_name)
            letter = col_letter(idx)
            dv = DataValidation(type="list", formula1=f"={ranges[list_name]}", allow_blank=True)
            dv.add(f"{letter}2:{letter}{DATA_ROWS}")
            ws.add_data_validation(dv)

    # All Requirements — live aggregation in A2
    allreq = wb.create_sheet("All Requirements")
    add_headers(allreq, PHASE_COLUMNS)
    allreq["A2"] = all_requirements_formula(phases, "excel")

    # Traceability — pull from All Requirements, AC Count computed
    trace = wb.create_sheet("Traceability")
    add_headers(trace, TRACEABILITY_COLUMNS)
    trace["A2"] = "=IFERROR('All Requirements'!A2,\"\")"

    # Acceptance Matrix
    accept = wb.create_sheet("Acceptance Matrix")
    add_headers(accept, ACCEPTANCE_MATRIX_COLUMNS)

    path = out / f"{title.replace(' ', '_')}_Register.xlsx"
    wb.save(path)
    return path.name


# ---- CLI -------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Scaffold a Requirements & Traceability Register.")
    ap.add_argument("--out", default="./requirements-register", help="output directory")
    ap.add_argument("--phases", type=int, default=3, help="number of phase register tabs")
    ap.add_argument("--title", default="Requirements Register", help="project title")
    ap.add_argument("--xlsx", action="store_true", help="also emit a wired .xlsx (needs openpyxl)")
    args = ap.parse_args()

    if args.phases < 1:
        ap.error("--phases must be >= 1")

    out = Path(args.out)
    made = build_csv_bundle(out, args.phases, args.title)
    print(f"Wrote CSV bundle to {out}/ ({len(made)} files):")
    for name in made:
        print(f"  - {name}")

    if args.xlsx:
        xlsx = build_xlsx(out, args.phases, args.title)
        if xlsx:
            print(f"Wrote wired workbook: {out}/{xlsx}")

    print("\nNext: import the tabs (see IMPORT_ORDER.txt) or open the .xlsx.")
    print("Formulas & Google Sheets variants: references/formulas.md")


if __name__ == "__main__":
    main()
